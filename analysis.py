import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl

st.set_page_config(page_title="TAT Calculator", page_icon="🚛", layout="centered")

st.title("🚛 Transport TAT Calculator")
st.markdown("Upload your Excel → Auto-calculate TAT columns in **HH:MM:SS** → Download as properly formatted Excel")
st.markdown("---")

# ── UPLOAD ──────────────────────────────────────────────────
uploaded = st.file_uploader("📂 Upload Excel File", type=["xlsx", "xls"])

if uploaded is None:
    st.info("👆 Please upload your transport Excel file to begin.")
    st.markdown("""
    **Columns calculated (all in HH:MM:SS format):**
    | Column | Formula |
    |---|---|
    | YI-GI | GateIn − YardIn |
    | GI-GW | GrossWeight − GateIn |
    | GW-TW | TareWeight − GrossWeight |
    | TW-GO | GateOut − TareWeight |
    | GI-GO | GateOut − GateIn |
    """)
    st.stop()

# ── LOAD FILE ────────────────────────────────────────────────
try:
    file_bytes = uploaded.read()
    df = pd.read_excel(io.BytesIO(file_bytes), keep_default_na=False, na_filter=False)
    df.columns = df.columns.str.strip()
    total_rows = len(df)
    st.success(f"✅ File loaded: **{total_rows} rows, {len(df.columns)} columns**")
except Exception as e:
    st.error(f"❌ Could not read file: {e}")
    st.stop()

with st.expander("🔍 Debug Info"):
    st.write(f"**Rows:** {len(df)} | **Columns:** {len(df.columns)}")
    st.write("**Column names:**", df.columns.tolist())
    st.dataframe(df.head(3), use_container_width=True)

st.markdown("---")

st.subheader("🔧 Map Your Columns")

all_cols = ["-- Not Available --"] + df.columns.tolist()

def auto_index(name):
    if name in all_cols:
        return all_cols.index(name)
    lower_map = {c.lower().strip(): i for i, c in enumerate(all_cols)}
    return lower_map.get(name.lower(), 0)

c1, c2 = st.columns(2)
with c1:
    col_yardin  = st.selectbox("YardIn (datetime)",      all_cols, index=auto_index("YardIn"))
    col_gatein  = st.selectbox("GateIn (datetime)",      all_cols, index=auto_index("GateIn"))
    col_grosswt = st.selectbox("GrossWeight (datetime)", all_cols, index=auto_index("GrossWeight"))
with c2:
    col_tarewt  = st.selectbox("TareWeight (datetime)",  all_cols, index=auto_index("TareWeight"))
    col_gateout = st.selectbox("GateOut (datetime)",     all_cols, index=auto_index("GateOut"))

st.markdown("---")

if st.button("⚙️ Calculate TAT", type="primary", use_container_width=True):

    result = df.copy()
    errors  = []
    calc_log = []

    def to_dt(col_name):
        try:
            s = result[col_name].replace("", pd.NaT)
            return pd.to_datetime(s, dayfirst=True, errors='coerce')
        except Exception as ex:
            errors.append(f"DateTime parse failed for '{col_name}': {ex}")
            return None

    # ── helper: seconds → HH:MM:SS string ────────────────
    def sec_to_hms(sec):
        if pd.isna(sec) or sec < 0:
            return ""
        sec = int(sec)
        h = sec // 3600
        m = (sec % 3600) // 60
        s = sec % 60
        return f"{h:02d}:{m:02d}:{s:02d}"

    # ── helper: diff → HH:MM:SS column ───────────────────
    def to_hhmmss(a, b, label):
        try:
            diff_sec = (b - a).dt.total_seconds()
            neg = int((diff_sec < 0).sum())
            if neg > 0:
                calc_log.append(f"⚠️ {label}: {neg} rows had negative time — set to blank")
            return diff_sec.apply(sec_to_hms)
        except Exception as ex:
            errors.append(f"Time diff failed for {label}: {ex}")
            return None

    # Parse datetimes
    dt_yardin  = to_dt(col_yardin)  if col_yardin  != "-- Not Available --" else None
    dt_gatein  = to_dt(col_gatein)  if col_gatein  != "-- Not Available --" else None
    dt_grosswt = to_dt(col_grosswt) if col_grosswt != "-- Not Available --" else None
    dt_tarewt  = to_dt(col_tarewt)  if col_tarewt  != "-- Not Available --" else None
    dt_gateout = to_dt(col_gateout) if col_gateout != "-- Not Available --" else None

    # Parse summary
    st.markdown("#### 📅 DateTime Parse Summary")
    cols5 = st.columns(5)
    for widget, label, dt_s, col_n in [
        (cols5[0], "YardIn",      dt_yardin,  col_yardin),
        (cols5[1], "GateIn",      dt_gatein,  col_gatein),
        (cols5[2], "GrossWeight", dt_grosswt, col_grosswt),
        (cols5[3], "TareWeight",  dt_tarewt,  col_tarewt),
        (cols5[4], "GateOut",     dt_gateout, col_gateout),
    ]:
        if dt_s is not None:
            widget.metric(label, f"{int(dt_s.notna().sum())}/{len(dt_s)}", f"← {col_n}")
        else:
            widget.metric(label, "Not mapped", "")

    st.markdown("---")
    st.markdown("#### ⚙️ Calculation Results (HH:MM:SS)")

    # Calculate all 5 TAT columns
    calcs = [
        ("YI-GI", dt_yardin,  dt_gatein,  "YardIn",      "GateIn"),
        ("GI-GW", dt_gatein,  dt_grosswt, "GateIn",      "GrossWeight"),
        ("GW-TW", dt_grosswt, dt_tarewt,  "GrossWeight", "TareWeight"),
        ("TW-GO", dt_tarewt,  dt_gateout, "TareWeight",  "GateOut"),
        ("GI-GO", dt_gatein,  dt_gateout, "GateIn",      "GateOut"),
    ]

    calculated_cols = []
    for col_name, dt_a, dt_b, from_lbl, to_lbl in calcs:
        if dt_a is not None and dt_b is not None:
            val = to_hhmmss(dt_a, dt_b, col_name)
            if val is not None:
                result[col_name] = val
                filled = int((val != "").sum())
                calculated_cols.append(col_name)
                st.success(f"✅ **{col_name}** ({to_lbl} − {from_lbl}) → {filled}/{len(result)} rows | Sample: `{val[val != ''].iloc[0] if filled > 0 else 'N/A'}`")
        else:
            errors.append(f"{col_name}: {from_lbl} or {to_lbl} not mapped")

    if calc_log:
        for msg in calc_log:
            st.warning(msg)
    if errors:
        st.markdown("---")
        st.warning("⚠️ Skipped:")
        for e in errors:
            st.write(f"- {e}")

    # Row count check
    st.markdown("---")
    rc1, rc2, rc3 = st.columns(3)
    rc1.metric("Uploaded rows", total_rows)
    rc2.metric("Output rows",   len(result))
    rc3.metric("Match?", "✅ YES" if total_rows == len(result) else "❌ MISMATCH")

    # Preview
    st.markdown("---")
    st.subheader(f"👁 Full Preview — All {len(result)} Rows")
    preview_cols = [c for c in [
        "Trip ID","Vehicle Number","Transporter Name",
        "YardIn","GateIn","GrossWeight","TareWeight","GateOut",
        "YI-GI","GI-GW","GW-TW","TW-GO","GI-GO"
    ] if c in result.columns]
    st.dataframe(result[preview_cols], use_container_width=True, height=420)

    # ── BUILD PROPERLY FORMATTED EXCEL ───────────────────
    st.markdown("---")
    st.subheader("⬇️ Download")

    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        result.to_excel(writer, index=False, sheet_name="TAT Result")
        wb = writer.book
        ws = writer.sheets["TAT Result"]

        # ── Date format for datetime columns ─────────────
        dt_col_names = [col_yardin, col_gatein, col_grosswt, col_tarewt, col_gateout]
        dt_col_names = [c for c in dt_col_names if c != "-- Not Available --"]

        # Map column name → Excel column letter
        col_letter_map = {}
        for idx, col_name in enumerate(result.columns, 1):
            col_letter_map[col_name] = get_column_letter(idx)

        # Apply DD-MM-YYYY HH:MM:SS format to all datetime columns
        date_fmt = "DD-MM-YYYY HH:MM:SS"
        for col_name in dt_col_names:
            if col_name in col_letter_map:
                col_letter = col_letter_map[col_name]
                for row_num in range(2, len(result) + 2):  # skip header row 1
                    cell = ws[f"{col_letter}{row_num}"]
                    # If cell has a real datetime value, apply date format
                    if cell.value and not isinstance(cell.value, str):
                        cell.number_format = date_fmt
                    elif isinstance(cell.value, str) and cell.value != "":
                        # Try to convert string back to datetime for proper Excel storage
                        try:
                            parsed = pd.to_datetime(cell.value, dayfirst=True)
                            cell.value = parsed.to_pydatetime()
                            cell.number_format = date_fmt
                        except:
                            pass  # keep as string if parse fails

        # Apply [HH]:MM:SS text format label to TAT columns (stored as text HH:MM:SS)
        tat_green_fill = PatternFill("solid", start_color="E2EFDA")
        tat_font       = Font(name="Arial", size=10, bold=True, color="375623")
        center_align   = Alignment(horizontal="center", vertical="center")

        for col_name in ["YI-GI","GI-GW","GW-TW","TW-GO","GI-GO"]:
            if col_name in col_letter_map:
                col_letter = col_letter_map[col_name]
                for row_num in range(2, len(result) + 2):
                    cell = ws[f"{col_letter}{row_num}"]
                    cell.fill      = tat_green_fill
                    cell.font      = tat_font
                    cell.alignment = center_align
                    # Store as text explicitly so Excel doesn't convert to serial number
                    cell.number_format = "@"

        # ── Header styling ────────────────────────────────
        header_fill     = PatternFill("solid", start_color="1F4E79")
        header_fill_tat = PatternFill("solid", start_color="375623")
        header_font     = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        thin_border     = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        tat_cols_set = {"YI-GI","GI-GW","GW-TW","TW-GO","GI-GO"}

        for idx, col_name in enumerate(result.columns, 1):
            cell = ws.cell(row=1, column=idx)
            cell.font      = header_font
            cell.fill      = header_fill_tat if col_name in tat_cols_set else header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = thin_border

        ws.row_dimensions[1].height = 30

        # ── Column widths ─────────────────────────────────
        for idx, col_name in enumerate(result.columns, 1):
            col_letter = get_column_letter(idx)
            if col_name in tat_cols_set:
                ws.column_dimensions[col_letter].width = 14
            elif col_name in dt_col_names:
                ws.column_dimensions[col_letter].width = 22
            else:
                ws.column_dimensions[col_letter].width = 18

        # ── Freeze header row ─────────────────────────────
        ws.freeze_panes = "A2"

        # ── Alternating row colors ────────────────────────
        row_fill_even = PatternFill("solid", start_color="EBF3FB")
        row_fill_odd  = PatternFill("solid", start_color="FFFFFF")
        data_font     = Font(name="Arial", size=9)

        for row_num in range(2, len(result) + 2):
            bg = row_fill_even if row_num % 2 == 0 else row_fill_odd
            for col_idx in range(1, len(result.columns) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                col_name = result.columns[col_idx - 1]
                if col_name not in tat_cols_set:
                    cell.fill = bg
                    cell.font = data_font
                cell.border    = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

    buf.seek(0)

    st.info(f"✅ Excel file ready — **{len(result)} rows** | Datetime columns formatted as `DD-MM-YYYY HH:MM:SS` | TAT columns as `HH:MM:SS` text")

    st.download_button(
        label=f"⬇️ Download Formatted Excel ({len(result)} rows)",
        data=buf,
        file_name="TAT_Calculated_Result.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary"
    )